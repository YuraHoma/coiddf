#!/usr/bin/env python3
"""Перетворює файл перекладача (.xlsx) назад у translations.csv.

Перекладачеві незручно працювати з CSV: роздільники, лапки, кодування.
Тому назовні йде звичайна книга Excel (scripts/i18n/xlsx-export.py),
а сюди повертається вона ж із правками. Цей скрипт складає з неї той
самий CSV, який уже вміє читати npm run i18n:import.

Використання:
    python3 scripts/i18n/xlsx-import.py <файл.xlsx>
    npm run i18n:import

Потрібен openpyxl:  pip install openpyxl
"""
import csv
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Потрібен openpyxl: pip install openpyxl")

OUT = os.path.join(os.path.dirname(__file__), "translations.csv")
SHEET = "Переклад"
# порядок стовпців у книзі перекладача
COL = {"num": 1, "area": 2, "field": 3, "source_uk": 4, "english": 5,
       "approved": 6, "note": 7, "key": 8, "source_hash": 9}
HEADER = ["key", "status", "source_uk", "previous_uk", "english",
          "approved", "note", "source_hash"]


def main():
    if len(sys.argv) < 2:
        sys.exit("Вкажіть файл: python3 scripts/i18n/xlsx-import.py <файл.xlsx>")
    src = sys.argv[1]
    wb = load_workbook(src, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"У книзі немає вкладки «{SHEET}». Це точно той файл?")
    ws = wb[SHEET]

    rows, skipped, approved = [], 0, 0
    for i in range(2, ws.max_row + 1):
        def cell(name):
            v = ws.cell(row=i, column=COL[name]).value
            return "" if v is None else str(v).strip()

        key = cell("key")
        if not key:
            skipped += 1
            continue
        if cell("approved").lower() in ("так", "yes", "y", "x", "+", "1", "true"):
            approved += 1
        rows.append({
            "key": key,
            "status": "",           # імпорт статус не читає, він рахує його з approved
            "source_uk": cell("source_uk"),
            "previous_uk": "",
            "english": cell("english"),
            "approved": cell("approved"),
            "note": cell("note"),
            "source_hash": cell("source_hash"),
        })

    with open(OUT, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADER)
        w.writeheader()
        w.writerows(rows)

    print(f"Прочитано рядків:      {len(rows)}")
    print(f"Позначено «готово»:    {approved}")
    if skipped:
        print(f"Пропущено без ключа:   {skipped}")
    print(f"\nЗаписано {OUT}")
    print("Далі: npm run i18n:import")


if __name__ == "__main__":
    main()
