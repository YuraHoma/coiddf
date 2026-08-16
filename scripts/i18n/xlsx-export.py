#!/usr/bin/env python3
"""Збирає зручну книгу Excel для перекладача з translations.csv.

Перекладачеві незручно працювати з CSV: роздільники, лапки, кодування,
однакові на вигляд технічні стовпці. Тому назовні йде книга з двох
вкладок — інструкція і сама таблиця, де редагувати треба лише жовтий
стовпець. Технічні позначки лишаються праворуч сірими: без них файл
не завантажиться назад.

Використання:
    npm run i18n:export          # спершу оновити CSV
    python3 scripts/i18n/xlsx-export.py [шлях-виводу.xlsx]

Потрібен openpyxl:  pip install openpyxl
"""
import csv, json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), 'translations.csv')
import sys
OUT = sys.argv[1] if len(sys.argv) > 1 else 'ICDF-pereklad.xlsx'

# ключ → людська назва місця на сайті
AREAS = {
    'site': 'Шапка, меню й футер (на всіх сторінках)',
    'home': 'Головна сторінка',
    'pronas': 'Сторінка «Про нас»',
    'team': 'Сторінка «Наша команда»',
    'partners': 'Сторінка «Партнери»',
    'proyekty': 'Сторінка «Проєкти» (заголовки розділу)',
    'novyny': 'Сторінка «Новини» (заголовки розділу)',
    'legal': 'Сторінка «Юридична інформація»',
    'policy': 'Сторінка «Політика конфіденційності»',
    'reports': 'Сторінка «Річні звіти»',
    'feedback': 'Сторінка «Контакти» і форма',
    'contacts': 'Контактні дані',
}
FIELD = {
    'title': 'повна назва',
    'body': 'текст сторінки',
    'heading': 'заголовок',
    'lead': 'підзаголовок',
    'eyebrow': 'мітка над заголовком',
    'label': 'підпис',
    'text': 'текст',
    'note': 'примітка',
    'value': 'значення',
    'name': "ім'я",
    'role': 'посада',
    'bio': 'опис людини',
    'headline': 'заголовок банера',
    'orgBanner': 'назва організації',
    'primaryLabel': 'кнопка',
    'secondaryLabel': 'кнопка',
    'buttonLabel': 'кнопка',
    'linkLabel': 'підпис посилання',
    'ctaLabel': 'кнопка',
    'detailsLabel': 'кнопка «Детальніше»',
    'newsHeading': 'заголовок блоку новин',
    'orgFull': 'повна назва організації',
    'orgShort': 'скорочена назва',
    'orgLine1': 'назва в шапці, рядок 1',
    'orgLine2': 'назва в шапці, рядок 2',
    'imageAlt': 'опис фото для незрячих',
    'alt': 'опис фото для незрячих',
}

def area_of(key):
    if key.startswith('data/'):
        top = key.split('/', 1)[1].split('.', 1)[0]
        return AREAS.get(top, top)
    if key.startswith('projects/'):
        slug = key.split('/', 1)[1].rsplit('.', 1)[0]
        return f'Проєкт: {slug}'
    if key.startswith('news/'):
        slug = key.split('/', 1)[1].rsplit('.', 1)[0]
        return f'Новина: {slug}'
    return key

def field_of(key):
    tail = re.sub(r'\[\d+\]$', '', key.rsplit('.', 1)[-1])
    return FIELD.get(tail, '')

rows = list(csv.DictReader(open(SRC, encoding='utf-8-sig')))

wb = Workbook()

# ---------------------------------------------------------------- Інструкція
ws = wb.active
ws.title = 'Як працювати'
ws.sheet_view.showGridLines = False
instr = [
    ('Переклад сайту ICDF — англійська версія', 'h1'),
    ('', ''),
    ('Що це за файл', 'h2'),
    ('Тут увесь текст, який видно на англійській версії сайту. Зараз він перекладений машиною, '
     'тому потребує вичитки. Кожен рядок — окремий шматок тексту: заголовок, абзац, підпис кнопки.', ''),
    ('', ''),
    ('Що робити', 'h2'),
    ('1. Перейдіть на вкладку «Переклад».', ''),
    ('2. Читайте український оригінал і англійський переклад поруч.', ''),
    ('3. Правте текст ЛИШЕ у стовпці «Англійською» (він виділений жовтим).', ''),
    ('4. У стовпці «Готово» поставте «так» навпроти кожного рядка, який ви перевірили — '
     'навіть якщо нічого не міняли.', ''),
    ('5. Якщо є питання або сумнів — напишіть у стовпці «Коментар».', ''),
    ('6. Збережіть файл і поверніть його розробнику.', ''),
    ('', ''),
    ('Важливо', 'h2'),
    ('• Не міняйте, не сортуйте і не видаляйте рядки — переклад привʼязаний до них по порядку.', ''),
    ('• Не чіпайте сірі стовпці праворуч: це технічні позначки, без них файл не завантажиться назад.', ''),
    ('• Рядок, позначений «так», більше ніколи не буде перезаписаний машиною. '
     'Непозначені рядки машина може оновити, якщо зміниться український текст.', ''),
    ('', ''),
    ('Про що варто памʼятати', 'h2'),
    ('• Тексти проєктів і новин («текст сторінки») працюють подвійно: перші два-три речення '
     'сайт бере в картку розділу, у результати пошуку і в попередній перегляд посилання. '
     'Тож початок тексту має читатися самостійно.', ''),
    ('• Назви проєктів довгі й офіційні — так і має бути. Для вкладки браузера сайт '
     'скорочує їх сам, окремо нічого писати не треба.', ''),
    ('• Довжина решти — орієнтуйтесь на український оригінал поруч.', ''),
    ('', ''),
    ('Питання — до Юрія.', 'note'),
]
styles = {
    'h1': (Font(name='Arial', size=16, bold=True, color='2E4620'), 28),
    'h2': (Font(name='Arial', size=12, bold=True, color='47632F'), 22),
    'note': (Font(name='Arial', size=11, italic=True, color='666666'), 18),
    '': (Font(name='Arial', size=11), 18),
}
for i, (text, kind) in enumerate(instr, start=1):
    c = ws.cell(row=i, column=1, value=text)
    font, height = styles[kind]
    c.font = font
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[i].height = height
ws.column_dimensions['A'].width = 110

# ---------------------------------------------------------------- Переклад
ws2 = wb.create_sheet('Переклад')
HEAD = ['№', 'Де на сайті', 'Що саме', 'Українською (оригінал)', 'Англійською — правте тут',
        'Готово', 'Коментар', 'ключ (не чіпати)', 'мітка (не чіпати)']
head_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
head_fill = PatternFill('solid', fgColor='47632F')
edit_fill = PatternFill('solid', fgColor='FFF6D5')
tech_fill = PatternFill('solid', fgColor='EFEFEF')
thin = Side(style='thin', color='D5D5D5')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for j, h in enumerate(HEAD, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    c.font = head_font
    c.fill = head_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws2.row_dimensions[1].height = 34

body_font = Font(name='Arial', size=11)
tech_font = Font(name='Arial', size=9, color='999999')
for i, r in enumerate(rows, start=2):
    vals = [i - 1, area_of(r['key']), field_of(r['key']), r['source_uk'], r['english'],
            r['approved'], r['note'], r['key'], r['source_hash']]
    for j, v in enumerate(vals, start=1):
        c = ws2.cell(row=i, column=j, value=v)
        c.font = tech_font if j >= 8 else body_font
        c.alignment = Alignment(wrap_text=(j in (2, 4, 5, 7)), vertical='top')
        c.border = border
        if j == 5:
            c.fill = edit_fill
        elif j >= 8:
            c.fill = tech_fill

widths = {1: 5, 2: 34, 3: 22, 4: 58, 5: 58, 6: 10, 7: 26, 8: 30, 9: 14}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

dv = DataValidation(type='list', formula1='"так,ні"', allow_blank=True, showDropDown=False)
dv.prompt = 'Поставте «так», коли рядок перевірено'
dv.promptTitle = 'Готово'
ws2.add_data_validation(dv)
dv.add(f'F2:F{len(rows) + 1}')

ws2.freeze_panes = 'D2'
ws2.auto_filter.ref = f'A1:I{len(rows) + 1}'

wb.save(OUT)
print('готово:', OUT, '| рядків:', len(rows))
